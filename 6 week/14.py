# Вася планирует карьеру и переезд.
# Для это составил таблицу, в которой для каждого региона записал зарплаты для разных интересные ему профессий.
# Таблица доступна по ссылке https://stepik.org/media/attachments/lesson/245267/salaries.xlsx.
# Выведите название региона с самой высокой медианной зарплатой (медианой называется элемент, стоящий в середине массива после его упорядочивания) и, через пробел, название профессии с самой высокой средней зарплатой по всем регионам.

import openpyxl

from statistics import mean

regions, proffs = {}, {}

wb = openpyxl.load_workbook('data/salaries.xlsx')
sheet = wb.active

for column in range(2, 8):
    for row in range(2, 9):
        salary = sheet.cell(row=row, column=column).value
        prof = sheet.cell(row=1, column=column).value
        city = sheet.cell(row=row, column=1).value

        if city not in regions:
            regions[city] = []
        regions[city].append(salary)

        if prof not in proffs:
            proffs[prof] = []
        proffs[prof].append(salary)

for reg in regions:
    regions[reg] = mean(regions[reg])
for prof in proffs:
    proffs[prof] = mean(proffs[prof])

print(sorted(regions.items(), key=lambda x: x[1])[-1][0],
      sorted(proffs.items(), key=lambda x: x[1])[-1][0], sep=' ')