# для решения этой задачи необходимо установить библиотеку openpyxl, скачать файл https://stepik.org/media/attachments/lesson/245266/tab.xlsx и создать в папке с этим файлом скрипт со следующем содержанием:

import openpyxl

wb = openpyxl.load_workbook('data/tab11.xlsx')
sh = wb.active
nmin = sh.cell(row=7, column=2).value
for rownum in range(8, 28):
    nmin = min(nmin, sh.cell(row=rownum, column=2).value)
print(nmin)