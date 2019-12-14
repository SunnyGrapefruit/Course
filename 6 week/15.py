# Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно, составив справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков, жиров и углеводов на 100 грамм продукта.
# Ему не удалось найти всю информацию, поэтому некоторые ячейки остались незаполненными (можно считать их значение равным нулю).
# Также он использовал какой-то странный офисный пакет и разделял целую и дробную часть чисел запятой.
# Таблица доступа по ссылке https://stepik.org/media/attachments/lesson/245290/trekking1.xlsx
#
# Вася хочет минимизировать вес продуктов и для этого брать самые калорийные продукты.
# Помогите ему и упорядочите продукты по убыванию калорийности.
# В случае, если продукты имеют одинаковую калорийность - упорядочите их по названию.
# В качестве ответа необходимо сдать названия продуктов, по одному в строке.


import openpyxl

spisok = {}

wb = openpyxl.load_workbook('data/trekking1.xlsx')
sheet = wb.active

for row in range(2, 39):
    product = sheet.cell(row=row, column=1).value
    price = sheet.cell(row=row, column=2).value

    if product not in spisok:
        spisok[product] = []
    spisok[product].append(price)

d = sorted(sorted(spisok.items(), key=lambda x: x[0]), key=lambda x: x[1], reverse=True)

for i in d:
    print(i[0])

# import operator
# for k, v in sorted(sorted(spisok.items(), key=operator.itemgetter(0)), key=operator.itemgetter(1), reverse=True):
#     print(k)
